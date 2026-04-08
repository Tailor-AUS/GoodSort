from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

H = Font(bold=True, size=11, name='Arial', color='FFFFFF')
HF = PatternFill('solid', fgColor='16A34A')
B = Font(color='0000FF', name='Arial', size=10)
K = Font(color='000000', name='Arial', size=10)
BF = Font(bold=True, name='Arial', size=10)
GF = PatternFill('solid', fgColor='F2F2F2')
YF = PatternFill('solid', fgColor='FEF3C7')
GN = PatternFill('solid', fgColor='DCFCE7')
RD = PatternFill('solid', fgColor='FEE2E2')
T = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))

def hdr(ws, cols):
    for j, h in enumerate(cols, 1):
        c = ws.cell(row=1, column=j, value=h)
        c.font = H; c.fill = HF; c.alignment = Alignment(horizontal='center', wrap_text=True); c.border = T

def c(ws, r, col, v, fmt='#,##0', f=None, fl=None):
    cl = ws.cell(row=r, column=col, value=v); cl.border = T; cl.number_format = fmt
    if f: cl.font = f
    if fl: cl.fill = fl
    return cl

def sec(ws, r, text):
    c(ws, r, 1, text, f=BF, fl=GF)
    for col in range(2, 10): c(ws, r, col, '', fl=GF)

# ═══════════════════════════════════════
# SHEET 1: KEY ASSUMPTIONS
# ═══════════════════════════════════════
ws = wb.active; ws.title = 'Key Numbers'
ws.column_dimensions['A'].width = 48
ws.column_dimensions['B'].width = 16
ws.column_dimensions['C'].width = 50
hdr(ws, ['Parameter', 'Value', 'Source / Notes'])

rows = [
    (None, 'YOUR REVENUE PER CONTAINER (you are platform + runner)'),
    ('CDS refund from COEX', 0.10, 'Claimed at depot on your Scheme ID'),
    ('COEX handling fee (depot rate SEQ)', 0.0768, 'COEX 2025 — registered CRP'),
    ('Scrap material value (blended avg)', 0.015, '55% AL, 30% PET, 15% glass'),
    ('TOTAL IN per container', '=B2+B3+B4', 'What COEX + scrap pays you'),
    (None, ''),
    (None, 'YOUR COSTS PER CONTAINER'),
    ('Sorter credit (5c)', 0.05, 'Sorting credit — NOT a CDS refund'),
    ('TOTAL OUT per container', '=B8', 'Only cost is sorter credit'),
    (None, ''),
    (None, 'NET MARGIN PER CONTAINER'),
    ('YOU KEEP per container', '=B5-B9', 'This is your real margin'),
    ('Margin as %', '=B11/B5', 'Of total revenue'),
    (None, ''),
    (None, 'COLLECTION RUN COSTS (MOOROOKA)'),
    ('Distance: house circuit + depot', 5, 'km — Moorooka loop + Yeerongpilly 1.5km'),
    ('Fuel cost per km', 0.21, 'ULP $2.30/L April 2026'),
    ('Fuel per run', '=B15*B16', ''),
    ('Time per run (minutes)', 40, 'Short circuit — you know the streets'),
    ('Time cost (@ $30/hr)', '=B18/60*30', 'Your opportunity cost'),
    ('TOTAL cost per run', '=B17+B19', 'Fuel + time'),
    (None, ''),
    (None, 'BREAK EVEN PER RUN'),
    ('Containers to cover FUEL only', '=CEILING(B17/B11,1)', 'Minimum viable run'),
    ('Containers to cover FUEL + TIME', '=CEILING(B20/B11,1)', 'Worth your time'),
    ('Per bin (fuel only, 5 bins)', '=CEILING(B22/5,1)', 'Each bin needs this many'),
    ('Per bin (fuel + time, 5 bins)', '=CEILING(B23/5,1)', 'Each bin needs this many'),
    (None, ''),
    (None, 'WHAT DOES $1000 SEED BUY?'),
    ('Setup cost (5 bins + QR + bags)', 70, ''),
    ('Remaining capital', '=1000-B28', ''),
    ('Collection runs this covers (fuel)', '=FLOOR(B29/B17,1)', 'Just fuel'),
    ('Containers those runs can process', '=B30*B22*5', 'At break-even volume'),
]

r = 2
for row in rows:
    if row[0] is None:
        if row[1]:
            sec(ws, r, row[1])
        r += 1
        continue
    c(ws, r, 1, row[0])
    cl = c(ws, r, 2, row[1], f=B if isinstance(row[1], (int, float)) else K)
    if isinstance(row[1], float) and 0 < row[1] < 1:
        cl.number_format = '$0.0000' if row[1] < 0.05 else '$0.00'
    elif isinstance(row[1], (int, float)) and row[1] >= 100:
        cl.number_format = '#,##0'
    elif isinstance(row[1], str) and '/' in str(row[1]):
        cl.number_format = '0.0%'
    elif isinstance(row[1], str) and row[1].startswith('='):
        if 'B11' in row[1] and 'B5' in row[1]:
            cl.number_format = '0.0%'
        else:
            cl.number_format = '$#,##0.00'
    c(ws, r, 3, row[2])
    r += 1

# Highlight key outputs
for hr in [5, 9, 11, 22, 23, 24, 25]:
    ws.cell(row=hr, column=1).fill = YF
    ws.cell(row=hr, column=2).fill = YF
    ws.cell(row=hr, column=3).fill = YF

# ═══════════════════════════════════════
# SHEET 2: HOW MANY BINS?
# ═══════════════════════════════════════
ws2 = wb.create_sheet('How Many Bins')
for col in range(1, 10):
    ws2.column_dimensions[get_column_letter(col)].width = 16
hdr(ws2, ['# Bins', 'Cans/Bin/Wk', 'Total/Week', 'Weekly Rev (kept)',
          'Weekly Fuel', 'Weekly Time Cost', 'Weekly Profit',
          'Monthly Profit', 'Annual Profit'])

# Model different bin counts at 50 containers/bin/week (conservative)
bin_counts = [5, 10, 15, 20, 30, 50, 75, 100, 150, 200, 300, 500]
cpb = 50  # conservative containers per bin per week

for i, bins in enumerate(bin_counts, 2):
    r = i
    c(ws2, r, 1, bins, f=B)
    c(ws2, r, 2, cpb, f=B)
    c(ws2, r, 3, f'=A{r}*B{r}', '#,##0')
    c(ws2, r, 4, f"=C{r}*'Key Numbers'!B11", '$#,##0')
    # Fuel: 1 run per 5 bins per week, 5km per run
    c(ws2, r, 5, f"=CEILING(A{r}/5,1)*'Key Numbers'!B17", '$#,##0')
    # Time: 40min per run
    c(ws2, r, 6, f"=CEILING(A{r}/5,1)*'Key Numbers'!B19", '$#,##0')
    c(ws2, r, 7, f'=D{r}-E{r}-F{r}', '$#,##0;($#,##0);-')
    c(ws2, r, 8, f'=G{r}*4.33', '$#,##0;($#,##0);-')
    c(ws2, r, 9, f'=G{r}*52', '$#,##0;($#,##0);-')

# Now same table at 100/bin/week (realistic)
r = len(bin_counts) + 3
c(ws2, r, 1, 'AT 100 CONTAINERS/BIN/WEEK (established)', f=BF, fl=GF)
for col in range(2, 10): c(ws2, r, col, '', fl=GF)
r += 1

for i, bins in enumerate(bin_counts):
    rr = r + i
    c(ws2, rr, 1, bins, f=B)
    c(ws2, rr, 2, 100, f=B)
    c(ws2, rr, 3, f'=A{rr}*B{rr}', '#,##0')
    c(ws2, rr, 4, f"=C{rr}*'Key Numbers'!B11", '$#,##0')
    c(ws2, rr, 5, f"=CEILING(A{rr}/5,1)*'Key Numbers'!B17", '$#,##0')
    c(ws2, rr, 6, f"=CEILING(A{rr}/5,1)*'Key Numbers'!B19", '$#,##0')
    c(ws2, rr, 7, f'=D{rr}-E{rr}-F{rr}', '$#,##0;($#,##0);-')
    c(ws2, rr, 8, f'=G{rr}*4.33', '$#,##0;($#,##0);-')
    c(ws2, rr, 9, f'=G{rr}*52', '$#,##0;($#,##0);-')

# ═══════════════════════════════════════
# SHEET 3: FIRST 6 MONTHS
# ═══════════════════════════════════════
ws3 = wb.create_sheet('First 6 Months')
for col in range(1, 12):
    ws3.column_dimensions[get_column_letter(col)].width = 14
hdr(ws3, ['Month', 'Bins', 'Cans/Bin/Wk', 'Total/Mo', 'COEX Revenue',
          'Scrap Revenue', 'Sorter Payouts', 'Fuel', 'Revenue Kept',
          'NET Profit', 'Cumulative'])

# Realistic: start 5 bins, grow to 15 by month 6
# Containers grow as word spreads
plan = [
    (1, 5, 25),   # Just placed, barely known
    (2, 5, 35),   # Neighbours starting to use
    (3, 8, 45),   # Added 3 more bins
    (4, 10, 55),  # Growing
    (5, 12, 65),  # More bins, better adoption
    (6, 15, 75),  # 15 bins, established
]

for i, (month, bins, cpw) in enumerate(plan, 2):
    r = i
    c(ws3, r, 1, month)
    c(ws3, r, 2, bins, f=B)
    c(ws3, r, 3, cpw, f=B)
    c(ws3, r, 4, f'=B{r}*C{r}*4.33', '#,##0')                    # Total/month
    c(ws3, r, 5, f"=D{r}*('Key Numbers'!B2+'Key Numbers'!B3)", '$#,##0')  # COEX revenue
    c(ws3, r, 6, f"=D{r}*'Key Numbers'!B4", '$#,##0')             # Scrap
    c(ws3, r, 7, f"=D{r}*'Key Numbers'!B8", '($#,##0)')           # Sorter payouts
    c(ws3, r, 8, f"=CEILING(B{r}/5,1)*4.33*'Key Numbers'!B17", '($#,##0)')  # Fuel
    c(ws3, r, 9, f'=E{r}+F{r}-G{r}', '$#,##0')                   # Revenue kept
    c(ws3, r, 10, f'=I{r}-H{r}', '$#,##0;($#,##0);-')            # NET
    if month == 1:
        c(ws3, r, 11, f'=J{r}-70', '$#,##0;($#,##0);-')          # Minus setup
    else:
        c(ws3, r, 11, f'=K{r-1}+J{r}', '$#,##0;($#,##0);-')

# Totals
r = 8
for col in range(1, 12):
    ws3.cell(row=r, column=col).fill = GF; ws3.cell(row=r, column=col).border = T
c(ws3, r, 1, 'TOTAL', f=BF)
for col in [4,5,6,7,8,9,10]:
    c(ws3, r, col, f'=SUM({get_column_letter(col)}2:{get_column_letter(col)}7)',
      '$#,##0;($#,##0);-', f=BF, fl=GF)

# ═══════════════════════════════════════
# SHEET 4: THE ANSWER
# ═══════════════════════════════════════
ws4 = wb.create_sheet('THE ANSWER')
ws4.column_dimensions['A'].width = 50
ws4.column_dimensions['B'].width = 18
ws4.column_dimensions['C'].width = 45
hdr(ws4, ['Question', 'Answer', 'How'])

answers = [
    (None, 'HOW MANY BINS TO START?'),
    ('Start with', 5, 'Moorooka — your street + 4 neighbours'),
    ('Cost to start', 70, '$10/bin + bags + QR stickers'),
    ('Wait until total across 5 bins hits', '=CEILING(\'Key Numbers\'!B17/\'Key Numbers\'!B11,1)', 'Then do first collection run'),
    ('Thats only', '=B4', 'containers to cover fuel'),
    (None, ''),
    (None, 'WHEN DOES IT MAKE REAL MONEY?'),
    ('At 5 bins × 50/wk you make', '=5*50*\'Key Numbers\'!B11*52', '/year (just from 5 bins!)'),
    ('At 15 bins × 75/wk you make', '=15*75*\'Key Numbers\'!B11*52', '/year'),
    ('At 50 bins × 100/wk you make', '=50*100*\'Key Numbers\'!B11*52', '/year'),
    ('At 100 bins × 100/wk you make', '=100*100*\'Key Numbers\'!B11*52', '/year'),
    (None, ''),
    (None, 'THE MAGIC NUMBER'),
    ('You keep per container', '=\'Key Numbers\'!B11', 'After sorter payout'),
    ('To make $1000/week you need', '=CEILING(1000/B14,1)', 'containers/week'),
    ('Thats how many bins @ 50/wk', '=CEILING(B15/50,1)', 'bins needed'),
    ('Thats how many bins @ 100/wk', '=CEILING(B15/100,1)', 'bins needed'),
    (None, ''),
    (None, 'TO MAKE $100K/YEAR'),
    ('Containers per week needed', '=CEILING(100000/52/\'Key Numbers\'!B11,1)', ''),
    ('Bins needed @ 50/wk', '=CEILING(B20/50,1)', ''),
    ('Bins needed @ 100/wk', '=CEILING(B20/100,1)', ''),
    ('Bins needed @ 150/wk', '=CEILING(B20/150,1)', ''),
    (None, ''),
    (None, 'BOTTOM LINE'),
    ('14c margin × volume = profit', '', 'No body corp fee needed'),
    ('5 bins pays for your time', '', 'From week 1'),
    ('50 bins replaces a salary', '', '$50-70k/year territory'),
    ('100+ bins = real business', '', '$100k+ with minimal overhead'),
]

r = 2
for row in answers:
    if row[0] is None:
        if row[1]:
            sec(ws4, r, row[1])
        r += 1
        continue
    c(ws4, r, 1, row[0])
    cl = c(ws4, r, 2, row[1], f=B if isinstance(row[1], (int, float)) else K)
    if isinstance(row[1], (int, float)) and row[1] >= 100:
        cl.number_format = '#,##0'
    elif isinstance(row[1], str) and row[1].startswith('='):
        if 'B11' in row[1] and '52' not in row[1]:
            cl.number_format = '$0.00'
        else:
            cl.number_format = '$#,##0'
    c(ws4, r, 3, row[2])
    r += 1

wb.save('docs/unit-economics-model.xlsx')
print('Created: docs/unit-economics-model.xlsx')
